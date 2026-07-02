import json
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SAMPLE_JSON = PROJECT_ROOT / "data" / "sample_cases.json"
DEFAULT_SYNTHETIC_EXCEL = PROJECT_ROOT.parent / "outputs" / "CS_2026_synthetic.xlsx"


def load_sample_cases(
    sample_json_path=DEFAULT_SAMPLE_JSON,
    synthetic_excel_path=DEFAULT_SYNTHETIC_EXCEL,
    max_cases: int = 5,
) -> List[Dict[str, str]]:
    excel_cases = load_synthetic_excel_cases(synthetic_excel_path, max_cases=max_cases)
    if excel_cases:
        return excel_cases

    path = Path(sample_json_path)
    if not path.exists():
        return []

    with path.open("r", encoding="utf-8") as file:
        cases = json.load(file)

    return [_normalize_case(case, index + 1) for index, case in enumerate(cases[:max_cases])]


def load_synthetic_excel_cases(excel_path, max_cases: int = 5) -> List[Dict[str, str]]:
    path = Path(excel_path)
    if not path.exists():
        return []

    workbook = load_workbook(path, data_only=True, read_only=True)
    if "2026" not in workbook.sheetnames:
        return []

    sheet = workbook["2026"]
    headers = _header_map(sheet)
    issue_col = _first_existing(headers, ["이슈"])
    response_col = _first_existing(headers, ["대응"])
    category_col = _first_existing(headers, ["_x0008_카테고리2026", "카테고리2026", "카테고리"])

    if not issue_col or not response_col or not category_col:
        return []

    cases = []
    for row in range(2, sheet.max_row + 1):
        issue = _cell_text(sheet.cell(row, issue_col).value)
        response = _cell_text(sheet.cell(row, response_col).value)
        category = _cell_text(sheet.cell(row, category_col).value)
        if not issue or not response or not category:
            continue

        cases.append(
            {
                "title": f"{category} 샘플 {len(cases) + 1}",
                "issue_text": issue,
                "response_text": response,
                "expected_category": category,
            }
        )
        if len(cases) >= max_cases:
            break

    return cases


def _header_map(sheet) -> Dict[str, int]:
    headers = {}
    for column in range(1, sheet.max_column + 1):
        value = _cell_text(sheet.cell(1, column).value)
        if value:
            headers[value] = column
    return headers


def _first_existing(headers: Dict[str, int], candidates: List[str]) -> Optional[int]:
    for candidate in candidates:
        if candidate in headers:
            return headers[candidate]
    return None


def _normalize_case(case: Dict[str, str], index: int) -> Dict[str, str]:
    expected = str(case.get("expected_category", "")).strip()
    return {
        "title": str(case.get("title") or f"{expected or '가상'} 샘플 {index}"),
        "issue_text": str(case.get("issue_text", "")).strip(),
        "response_text": str(case.get("response_text", "")).strip(),
        "expected_category": expected,
    }


def _cell_text(value) -> str:
    return "" if value is None else str(value).strip()
